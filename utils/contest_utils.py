from datetime import datetime
from typing import Optional, List, Dict, Tuple
from services.database import db, contest_participations_col
import os
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from bson import ObjectId
import logging
import io
import openpyxl.drawing.image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D

logger = logging.getLogger(__name__)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

async def save_contest_participation(
    contest_id: str,
    contest_name: str,
    date: str,
    level: str,
    teacher_name: str,
    nomination: str,
    participation_form: str,
    participant_type: str,
    student_name: Optional[str],
    group: Optional[str],
    result: str,
    confirmation_files: List[str],
    user_id: int
) -> None:
    """
    Сохраняет участие в конкурсе (отдельная запись для каждого участника)
    """
    try:
        # Инициализируем confirmation_files, если он None
        if confirmation_files is None:
            confirmation_files = []
            
        # Преобразуем confirmation_files в список, если это строка
        if isinstance(confirmation_files, str):
            confirmation_files = [confirmation_files]
            
        logger.info(f"Сохранение участия в конкурсе. User ID: {user_id}, Files: {confirmation_files}")
        
        contest_participations_col.insert_one({
            "contest_id": ObjectId(contest_id),
            "contest_name": contest_name,
            "date": date,
            "level": level,
            "teacher_name": teacher_name,
            "nomination": nomination,
            "participation_form": participation_form,
            "participant_type": participant_type,
            "student_name": student_name,
            "group": group,
            "result": result,
            "confirmation_files": confirmation_files,
            "user_id": user_id,
            "created_at": datetime.now()
        })
        logger.info(f"Успешно сохранено участие в конкурсе для пользователя {user_id}")
    except Exception as e:
        logger.error(f"Ошибка при сохранении участия в конкурсе: {e}", exc_info=True)
        raise

async def generate_contest_report(month: int, year: int) -> Tuple[List[Dict], Dict[str, list]]:
    """
    Генерирует отчет по конкурсам за указанный месяц и год.
    Возвращает список данных и словарь с изображениями в формате base64.
    """
    logger.info(f"Начинаем генерацию отчета за {month}.{year}")
    
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    records = list(db.contest_participations.find({
        "created_at": {"$gte": start_date, "$lt": end_date}
    }))
    
    logger.info(f"Найдено записей в базе данных: {len(records)}")

    if not records:
        return [], {}

    data = []
    images = {}

    for doc in records:
        confirmation_files = doc.get("confirmation_files", [])
        logger.debug(f"confirmation_files: {confirmation_files}")
        
        files_info = []
        for file_info in confirmation_files:
            if isinstance(file_info, dict):
                saved_name = file_info.get("saved_name")
                original_name = file_info.get("original_name")
                file_id = file_info.get("file_id")
            else:
                saved_name = file_info  # Убираем добавление .jpg
                original_name = saved_name
                file_id = file_info
                
            logger.debug(f"Обработка файла: saved_name={saved_name}, original_name={original_name}, file_id={file_id}")
            
            # Проверяем, есть ли расширение в имени файла
            if not os.path.splitext(saved_name)[1]:
                saved_name = f"{saved_name}.jpg"
            
            file_path = os.path.join(UPLOAD_FOLDER, saved_name)
            if os.path.exists(file_path):
                logger.debug(f"Файл найден: {file_path}")
                with open(file_path, "rb") as img_file:
                    image_data = base64.b64encode(img_file.read()).decode()
                    images[file_id] = [image_data]
                files_info.append({
                    "file_id": file_id,
                    "name": original_name
                })
            else:
                logger.warning(f"Файл не найден: {file_path}")
                
        data.append({
            "Название конкурса": doc.get("contest_name", ""),
            "Дата": doc.get("date", ""),
            "Уровень конкурса": doc.get("level", ""),
            "ФИО преподавателя": doc.get("teacher_name", ""),
            "Номинация": doc.get("nomination", ""),
            "Форма участия": doc.get("participation_form", ""),
            "Участник": doc.get("participant_type", ""),
            "ФИО студента": doc.get("student_name", ""),
            "Группа": doc.get("group", ""),
            "Результат": doc.get("result", ""),
            "Файлы": ", ".join(f_info["name"] for f_info in files_info),
            "confirmation_files": [f_info["file_id"] for f_info in files_info]
        })
        
    logger.info(f"Подготовлено данных: {len(data)}, изображений: {len(images)}")
    return data, images

async def create_contest_excel_report(data: List[Dict], images: Dict[str, list]) -> bytes:
    """
    Создание Excel-отчета по конкурсам с данными и изображениями
    
    Args:
        data: Список словарей с данными
        images: Словарь с изображениями в формате base64
    
    Returns:
        bytes: Бинарные данные Excel-файла
    """
    logger.info(f"Начинаем создание отчета по конкурсам. Количество записей: {len(data)}")
    
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по конкурсам"
    
    # Определяем заголовки
    headers = [
        "Название конкурса", "Дата", "Уровень конкурса", "ФИО преподавателя", "Номинация",
        "Форма участия", "Участник", "ФИО студента", "Группа", "Результат", "Файлы"
    ]
    
    # Настраиваем стили
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = wrap_alignment
    
    # Константы для форматирования
    ROW_HEIGHT = 150
    IMG_WIDTH = 150
    MAX_IMG_HEIGHT = 120
    
    # Создаем временную директорию для изображений
    temp_dir = os.path.join(UPLOAD_FOLDER, "temp")
    os.makedirs(temp_dir, exist_ok=True)
    
    try:
        # Записываем данные
        for row_idx, row_data in enumerate(data, 2):
            logger.debug(f"Обработка строки {row_idx}")
            
            # Устанавливаем высоту строки
            ws.row_dimensions[row_idx].height = ROW_HEIGHT
            
            # Записываем основные данные
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                cell.border = thin_border
                cell.alignment = wrap_alignment
            
            # Добавляем изображения
            confirmation_files = row_data.get("confirmation_files", [])
            img_col = len(headers) + 1
            
            
            for file_id in confirmation_files:
                if file_id in images:
                    for img_base64 in images[file_id]:
                        try:
                            
                            # Декодируем base64 в бинарные данные
                            img_data = base64.b64decode(img_base64)
                            logger.debug(f"Изображение {file_id} успешно декодировано из base64")
                            
                            # Сохраняем во временный файл
                            temp_img_path = os.path.join(temp_dir, f"temp_{file_id}.png")
                            with open(temp_img_path, "wb") as f:
                                f.write(img_data)
                            logger.debug(f"Изображение {file_id} сохранено во временный файл: {temp_img_path}")
                            
                            # Открываем изображение с помощью PIL
                            with Image.open(temp_img_path) as pil_img:
                                # Получаем размеры изображения
                                width, height = pil_img.size
                                logger.debug(f"Размеры изображения {file_id}: {width}x{height}")
                                
                                # Вычисляем новые размеры с сохранением пропорций
                                if width > height:
                                    new_width = min(IMG_WIDTH, width)
                                    new_height = int((height * new_width) / width)
                                else:
                                    new_height = min(MAX_IMG_HEIGHT, height)
                                    new_width = int((width * new_height) / height)
                                logger.debug(f"Новые размеры изображения {file_id}: {new_width}x{new_height}")
                                
                                # Изменяем размер изображения
                                pil_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                                pil_img.save(temp_img_path, format='PNG')
                                logger.debug(f"Изображение {file_id} изменено и сохранено")
                            
                            # Создаем объект изображения для openpyxl
                            img = openpyxl.drawing.image.Image(temp_img_path)
                            logger.debug(f"Создан объект изображения для Excel: {file_id}")
                            
                            # Добавляем изображение в ячейку
                            cell = ws.cell(row=row_idx, column=img_col)
                            cell.value = None  # Очищаем ячейку
                            logger.debug(f"Очищена ячейка {get_column_letter(img_col)}{row_idx}")
                            
                            # Устанавливаем размер ячейки
                            ws.column_dimensions[get_column_letter(img_col)].width = new_width * 0.14
                            logger.debug(f"Установлена ширина столбца {get_column_letter(img_col)}: {new_width * 0.14}")
                            
                            # Добавляем изображение
                            ws.add_image(img, f"{get_column_letter(img_col)}{row_idx}")
                            
                            img_col += 1
                            
                        except Exception as e:
                            logger.error(f"Ошибка при обработке изображения {file_id}: {str(e)}", exc_info=True)
                            continue
                else:
                    logger.warning(f"Изображение {file_id} не найдено в словаре images")
        
        # Настраиваем ширину столбцов с данными
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            
            # Проверяем длину заголовка
            cell = ws[f"{column}1"]
            max_length = len(str(cell.value))
            
            # Проверяем длину данных
            for row in range(2, len(data) + 2):
                cell = ws[f"{column}{row}"]
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Устанавливаем ширину с небольшим отступом
            adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
            ws.column_dimensions[column].width = adjusted_width
        
        logger.info("Сохранение отчета...")
        # Сохраняем файл в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Отчет успешно создан")
        return output.getvalue()
    
    except Exception as e:
        logger.error(f"Ошибка при создании отчета: {str(e)}")
        raise
    
    finally:
        # Удаляем временную директорию со всеми файлами
        import shutil
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            logger.debug("Временная директория удалена")

async def create_contest_html_report(data: List[Dict], images: Dict[str, list]) -> str:
    """
    Создание HTML-отчета по конкурсам с данными и изображениями
    
    Args:
        data: Список словарей с данными
        images: Словарь с изображениями в формате base64
    
    Returns:
        str: HTML-код отчета
    """
    logger.info(f"Начинаем создание HTML-отчета по конкурсам. Количество записей: {len(data)}")
    
    # Создаем HTML-шаблон с поддержкой сортировки и фильтрации
    html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Отчет по конкурсам</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
            }
            .controls {
                margin: 20px 0;
                padding: 15px;
                background-color: #f5f5f5;
                border-radius: 5px;
            }
            .filter-group {
                margin: 10px 0;
            }
            .filter-group label {
                display: inline-block;
                width: 150px;
                margin-right: 10px;
            }
            .filter-group input, .filter-group select {
                padding: 5px;
                margin: 5px 0;
                width: 200px;
            }
            .buttons {
                margin: 15px 0;
            }
            .buttons button {
                padding: 8px 15px;
                margin-right: 10px;
                cursor: pointer;
                background-color: #D7E4BC;
                border: 1px solid #999;
                border-radius: 3px;
            }
            .buttons button:hover {
                background-color: #C5D4A9;
            }
            table {
                border-collapse: collapse;
                width: 100%;
                margin: 20px 0;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
                vertical-align: top;
            }
            th {
                background-color: #D7E4BC;
                cursor: pointer;
                position: relative;
            }
            th:hover {
                background-color: #C5D4A9;
            }
            th::after {
                content: '↕';
                position: absolute;
                right: 8px;
                color: #666;
            }
            th.asc::after {
                content: '↑';
            }
            th.desc::after {
                content: '↓';
            }
            .images-cell {
                display: flex;
                flex-wrap: wrap;
                gap: 10px;
            }
            .image-container {
                max-width: 150px;
                margin: 5px;
            }
            .image-container img {
                max-width: 100%;
                height: auto;
            }
            .hidden {
                display: none;
            }
            .stats {
                margin: 10px 0;
                padding: 10px;
                background-color: #f0f0f0;
                border-radius: 3px;
            }
            @media print {
                .controls, .buttons {
                    display: none;
                }
                table {
                    width: 100%;
                }
                th, td {
                    border: 1px solid #000;
                }
                .images-cell {
                    page-break-inside: avoid;
                }
                .image-container {
                    max-width: 200px;
                }
            }
        </style>
        <script>
            let originalData = [];
            let currentFilters = {};
            let currentSort = { column: null, direction: 'asc' };
            
            function initializeData() {
                const table = document.getElementById('contestTable');
                const rows = Array.from(table.getElementsByTagName('tbody')[0].getElementsByTagName('tr'));
                originalData = rows.map(row => {
                    const cells = Array.from(row.getElementsByTagName('td'));
                    const imagesCell = cells[cells.length - 1]; // Последняя ячейка содержит изображения
                    return {
                        element: row,
                        data: cells.map(cell => cell.textContent.trim()),
                        images: imagesCell.classList.contains('images-cell') ? 
                            imagesCell.getElementsByTagName('img').length : 0
                    };
                });
                updateStats();
            }
            
            function updateStats() {
                const total = originalData.length;
                const visible = document.querySelectorAll('#contestTable tbody tr:not(.hidden)').length;
                document.getElementById('stats').textContent = `Показано ${visible} из ${total} записей`;
            }
            
            function applyFilters() {
                const filters = {};
                document.querySelectorAll('.filter-group input, .filter-group select').forEach(input => {
                    if (input.value) {
                        filters[input.name] = input.value.toLowerCase();
                    }
                });
                currentFilters = filters;
                filterAndSort();
            }
            
            function filterAndSort() {
                const tbody = document.querySelector('#contestTable tbody');
                const rows = Array.from(tbody.getElementsByTagName('tr'));
                
                // Сначала применяем фильтры
                rows.forEach(row => {
                    let visible = true;
                    const cells = Array.from(row.getElementsByTagName('td'));
                    
                    Object.entries(currentFilters).forEach(([key, value]) => {
                        const index = parseInt(key);
                        if (!cells[index].textContent.trim().toLowerCase().includes(value.toLowerCase())) {
                            visible = false;
                        }
                    });
                    
                    row.classList.toggle('hidden', !visible);
                });
                
                // Затем применяем сортировку
                if (currentSort.column !== null) {
                    const visibleRows = rows.filter(row => !row.classList.contains('hidden'));
                    const sortIndex = currentSort.column;
                    const direction = currentSort.direction;
                    
                    visibleRows.sort((a, b) => {
                        const aCells = Array.from(a.getElementsByTagName('td'));
                        const bCells = Array.from(b.getElementsByTagName('td'));
                        
                        if (sortIndex === 10) { // Сортировка по количеству изображений
                            const aImages = aCells[sortIndex].getElementsByTagName('img').length;
                            const bImages = bCells[sortIndex].getElementsByTagName('img').length;
                            return direction === 'asc' ? aImages - bImages : bImages - aImages;
                        } else {
                            const aValue = aCells[sortIndex].textContent.trim();
                            const bValue = bCells[sortIndex].textContent.trim();
                            return direction === 'asc' ? 
                                aValue.localeCompare(bValue) : 
                                bValue.localeCompare(aValue);
                        }
                    });
                    
                    // Перемещаем отсортированные строки в DOM
                    visibleRows.forEach(row => tbody.appendChild(row));
                }
                
                updateStats();
            }
            
            function sortTable(n) {
                const headers = document.getElementsByTagName('th');
                
                // Сбрасываем классы сортировки со всех заголовков
                Array.from(headers).forEach(header => {
                    header.classList.remove('asc', 'desc');
                });
                
                // Определяем направление сортировки
                if (currentSort.column === n) {
                    currentSort.direction = currentSort.direction === 'asc' ? 'desc' : 'asc';
                } else {
                    currentSort.column = n;
                    currentSort.direction = 'asc';
                }
                
                // Добавляем класс сортировки к текущему заголовку
                headers[n].classList.add(currentSort.direction);
                
                filterAndSort();
            }
            
            function resetFilters() {
                document.querySelectorAll('.filter-group input, .filter-group select').forEach(input => {
                    input.value = '';
                });
                currentFilters = {};
                filterAndSort();
            }
            
            function resetSort() {
                currentSort = { column: null, direction: 'asc' };
                document.querySelectorAll('th').forEach(th => {
                    th.classList.remove('asc', 'desc');
                });
                filterAndSort();
            }
            
            function showAll() {
                originalData.forEach(item => {
                    item.element.classList.remove('hidden');
                });
                updateStats();
            }
            
            // Инициализация при загрузке страницы
            window.onload = function() {
                initializeData();
            };
        </script>
    </head>
    <body>
        <h1>Отчет по конкурсам</h1>
        
        <div class="controls">
            <div class="filter-group">
                <label for="filter0">Название конкурса:</label>
                <input type="text" id="filter0" name="0" oninput="applyFilters()">
            </div>
            <div class="filter-group">
                <label for="filter1">Дата:</label>
                <input type="text" id="filter1" name="1" oninput="applyFilters()">
            </div>
            <div class="filter-group">
                <label for="filter2">Уровень конкурса:</label>
                <input type="text" id="filter2" name="2" oninput="applyFilters()">
            </div>
            <div class="filter-group">
                <label for="filter3">ФИО преподавателя:</label>
                <input type="text" id="filter3" name="3" oninput="applyFilters()">
            </div>
            <div class="filter-group">
                <label for="filter4">Номинация:</label>
                <input type="text" id="filter4" name="4" oninput="applyFilters()">
            </div>
            <div class="filter-group">
                <label for="filter5">Форма участия:</label>
                <select id="filter5" name="5" onchange="applyFilters()">
                    <option value="">Все</option>
                    <option value="Очная">Очная</option>
                    <option value="Заочная">Заочная</option>
                </select>
            </div>
            <div class="filter-group">
                <label for="filter6">Участник:</label>
                <select id="filter6" name="6" onchange="applyFilters()">
                    <option value="">Все</option>
                    <option value="Преподаватель">Преподаватель</option>
                    <option value="Студент">Студент</option>
                </select>
            </div>
            <div class="filter-group">
                <label for="filter7">ФИО студента:</label>
                <input type="text" id="filter7" name="7" oninput="applyFilters()">
            </div>
            <div class="filter-group">
                <label for="filter8">Группа:</label>
                <input type="text" id="filter8" name="8" oninput="applyFilters()">
            </div>
            <div class="filter-group">
                <label for="filter9">Результат:</label>
                <input type="text" id="filter9" name="9" oninput="applyFilters()">
            </div>
        </div>
        
        <div class="buttons">
            <button onclick="showAll()">Показать все записи</button>
            <button onclick="resetFilters()">Сбросить фильтры</button>
            <button onclick="resetSort()">Сбросить сортировку</button>
            <button onclick="window.print()">Печать</button>
        </div>
        
        <div class="stats" id="stats"></div>
        
        <table id="contestTable">
            <thead>
                <tr>
                    <th onclick="sortTable(0)">Название конкурса</th>
                    <th onclick="sortTable(1)">Дата</th>
                    <th onclick="sortTable(2)">Уровень конкурса</th>
                    <th onclick="sortTable(3)">ФИО преподавателя</th>
                    <th onclick="sortTable(4)">Номинация</th>
                    <th onclick="sortTable(5)">Форма участия</th>
                    <th onclick="sortTable(6)">Участник</th>
                    <th onclick="sortTable(7)">ФИО студента</th>
                    <th onclick="sortTable(8)">Группа</th>
                    <th onclick="sortTable(9)">Результат</th>
                    <th onclick="sortTable(10)">Фото</th>
                </tr>
            </thead>
            <tbody>
    """
    
    # Добавляем строки с данными
    for row_data in data:
        html_template += "<tr>"
        
        # Добавляем основные данные
        for header in ["Название конкурса", "Дата", "Уровень конкурса", "ФИО преподавателя", 
                      "Номинация", "Форма участия", "Участник", "ФИО студента", "Группа", "Результат"]:
            html_template += f"<td>{row_data.get(header, '')}</td>"
        
        # Добавляем ячейку с изображениями
        html_template += '<td class="images-cell">'
        confirmation_files = row_data.get("confirmation_files", [])
        for file_id in confirmation_files:
            if file_id in images:
                for img_base64 in images[file_id]:
                    html_template += f'''
                        <div class="image-container">
                            <img src="data:image/jpeg;base64,{img_base64}" alt="Фото подтверждения">
                        </div>
                    '''
        html_template += "</td>"
        
        html_template += "</tr>"
    
    # Завершаем HTML-шаблон
    html_template += """
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return html_template 