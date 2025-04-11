from datetime import datetime
from typing import Optional, List, Tuple, Dict
from services.database import db
import os
from io import BytesIO
import base64
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from PIL import Image
from config import logger

# Константа для пути к папке с загруженными файлами
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

async def save_self_assessment(
    user_id: int,
    event_type: str,
    event_name: str,
    description: str,
    result: str,
    social_media_link: Optional[str],
    confirmation_files: List[str],  # Список ID файлов подтверждения
    contest_id: Optional[str] = None
) -> None:
    """
    Сохранение данных самообследования в базу данных
    
    Args:
        user_id: ID пользователя
        event_type: Тип мероприятия
        event_name: Название мероприятия
        description: Описание мероприятия
        result: Результат участия
        social_media_link: Ссылка на публикацию в соцсетях
        confirmation_files: Список ID файлов подтверждения
        contest_id: ID конкурса (если применимо)
    """
    # Получаем информацию о пользователе
    user = db.users.find_one({"telegram_id": user_id})
    user_name = user.get("full_name", "Неизвестный пользователь") if user else "Неизвестный пользователь"
    
    db.self_assessments.insert_one({
        "user_id": user_id,
        "user_name": user_name,  # Сохраняем имя пользователя
        "event_type": event_type,
        "event_name": event_name,
        "description": description,
        "result": result,
        "social_media_link": social_media_link,
        "confirmation_files": confirmation_files,  # Сохраняем список ID файлов
        "contest_id": contest_id,
        "created_at": datetime.now()
    })

async def get_contests_by_type(event_type: str = None) -> list:
    """
    Получение списка конкурсов
    
    Args:
        event_type: Тип мероприятия (опционально). Если не указан, возвращаются все конкурсы
    
    Returns:
        list: Список конкурсов
    """
    if event_type:
        cursor = db.contests.find({"type": event_type})
    else:
        cursor = db.contests.find({})
    
    contests = list(cursor)
    return contests

async def generate_monthly_report(month: int, year: int) -> Tuple[List[Dict], Dict[str, list]]:
    """
    Генерирует отчет за указанный месяц и год.
    Возвращает список данных и словарь с изображениями в формате base64.
    """
    # Получаем данные за указанный месяц
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    # Получаем все записи за период
    records = list(db.self_assessments.find({
        "created_at": {
            "$gte": start_date,
            "$lt": end_date
        }
    }))
    
    if not records:
        return [], {}
    
    # Подготавливаем данные
    data = []
    images = {}
    
    for doc in records:
        # Получаем информацию о файлах подтверждения
        confirmation_files = doc.get("confirmation_files", [])
        files_info = []
        
        for file_info in confirmation_files:
            if isinstance(file_info, dict):
                # Новый формат (словарь с информацией о файле)
                saved_name = file_info.get("saved_name")
                original_name = file_info.get("original_name")
                file_id = file_info.get("file_id")
            else:
                # Старый формат (только file_id)
                saved_name = f"{file_info}.jpg"  # Предполагаем, что это фото
                original_name = saved_name
                file_id = file_info
            
            # Ищем файл в папке uploads
            file_path = os.path.join(UPLOAD_FOLDER, saved_name)
            if os.path.exists(file_path):
                # Конвертируем изображение в base64
                with open(file_path, "rb") as img_file:
                    image_data = base64.b64encode(img_file.read()).decode()
                    images[file_id] = [image_data]  # Сохраняем как список для совместимости
                files_info.append({
                    "file_id": file_id,
                    "name": original_name
                })
        
        # Добавляем запись в данные
        data.append({
            "ФИО": doc.get("user_name", "Неизвестный пользователь"),
            "Тип мероприятия": doc.get("event_type", ""),
            "Название мероприятия": doc.get("event_name", ""),
            "Описание": doc.get("description", ""),
            "Результат": doc.get("result", ""),
            "Ссылка на соц. сети": doc.get("social_media_link", ""),
            "Дата": doc.get("created_at", datetime.now()).strftime("%d.%m.%Y"),
            "Файлы": ", ".join(f_info["name"] for f_info in files_info),
            "confirmation_files": [f_info["file_id"] for f_info in files_info]
        })
    
    return data, images

async def get_activity_types() -> list:
    """
    Получение списка типов мероприятий из базы данных
    
    Returns:
        list: Список типов мероприятий в формате [(code, description), ...]
    """
    activity_types = db.activity_types.find_one({}) or {"types": []}
    return [(act.split(". ", 1)[0], act) for act in activity_types["types"]]

async def create_excel_report(data: List[Dict], images: Dict[str, list]) -> bytes:
    """
    Создание Excel-отчета с данными и изображениями
    
    Args:
        data: Список словарей с данными
        images: Словарь с изображениями в формате base64
    
    Returns:
        bytes: Бинарные данные Excel-файла
    """
    logger.info(f"Начинаем создание отчета. Количество записей: {len(data)}")
    
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    
    # Определяем заголовки
    headers = ["ФИО", "Тип мероприятия", "Название мероприятия", "Описание", 
              "Результат", "Ссылка на соц. сети", "Дата"]
    
    # Настраиваем стили
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = wrap_alignment
    
    # Максимальная высота строки (для изображений)
    ROW_HEIGHT = 150
    # Ширина изображения
    IMG_WIDTH = 150
    
    # Создаем временную директорию для изображений
    temp_dir = os.path.join(UPLOAD_FOLDER, "temp")
    os.makedirs(temp_dir, exist_ok=True)
    
    try:
        # Записываем данные
        for row_idx, row_data in enumerate(data, 2):
            
            # Устанавливаем высоту строки
            ws.row_dimensions[row_idx].height = ROW_HEIGHT
            
            # Записываем основные данные
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                cell.border = thin_border
                cell.alignment = wrap_alignment
            
            # Добавляем изображения
            confirmation_files = row_data.get("confirmation_files", [])
            img_col = len(headers) + 1
            
 
            
            for file_id in confirmation_files:
                if file_id in images:
                    for img_base64 in images[file_id]:
                        try:
                            # Декодируем base64 в бинарные данные
                            img_data = base64.b64decode(img_base64)
                            img_io = io.BytesIO(img_data)
                            
                            # Открываем изображение с помощью PIL для изменения размера
                            with Image.open(img_io) as pil_img:
                                # Получаем размеры изображения
                                width, height = pil_img.size
                                
                                # Вычисляем новую высоту, сохраняя пропорции
                                new_width = IMG_WIDTH
                                new_height = int((height * IMG_WIDTH) / width)
                                
                                # Если высота больше максимальной, пересчитываем ширину
                                if new_height > ROW_HEIGHT:
                                    new_width = int((width * ROW_HEIGHT) / height)
                                    new_height = ROW_HEIGHT
                                
                          
                                
                                # Изменяем размер изображения
                                pil_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                                
                                # Сохраняем измененное изображение во временный файл
                                temp_img_path = os.path.join(temp_dir, f"temp_{file_id}.png")
                                pil_img.save(temp_img_path, format='PNG')
                                
                                # Создаем объект изображения для openpyxl
                                import openpyxl.drawing.image
                                img = openpyxl.drawing.image.Image(temp_img_path)
                                
                                # Устанавливаем позицию изображения
                                # Используем абсолютное позиционирование
                                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                                from openpyxl.utils.units import pixels_to_EMU
                                from openpyxl.drawing.xdr import XDRPositiveSize2D
                                
                                # Создаем маркер для позиции
                                marker = AnchorMarker(col=img_col - 1, colOff=0, row=row_idx - 1, rowOff=0)
                                
                                # Создаем объект размера для изображения
                                size = XDRPositiveSize2D(cx=pixels_to_EMU(new_width), cy=pixels_to_EMU(new_height))
                                
                                # Создаем якорь с абсолютным позиционированием
                                img.anchor = OneCellAnchor(_from=marker, ext=size)
                                
                                # Добавляем изображение в лист
                                ws.add_image(img)
                                
                                # Устанавливаем ширину столбца
                                column_width = new_width * 0.14  # Примерное соотношение
                                ws.column_dimensions[get_column_letter(img_col)].width = column_width
                                
                                
                                img_col += 1
                        except Exception as e:
                            logger.error(f"Ошибка при обработке изображения: {str(e)}")
                            continue
        
        # Настраиваем ширину столбцов с данными
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            
            # Проверяем длину заголовка
            cell = ws[f"{column}1"]
            max_length = len(str(cell.value))
            
            # Проверяем длину данных
            for row in range(2, len(data) + 2):
                cell = ws[f"{column}{row}"]
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Устанавливаем ширину с небольшим отступом
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        logger.info("Сохранение отчета...")
        # Сохраняем файл в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Отчет успешно создан")
        return output.getvalue()
    
    finally:
        # Удаляем временную директорию со всеми файлами
        import shutil
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)